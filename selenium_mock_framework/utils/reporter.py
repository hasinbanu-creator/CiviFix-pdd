import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill

class Reporter:
    def __init__(self):
        self.reports_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "reports")
        self.excel_dir = os.path.join(self.reports_dir, "excel")
        self.html_dir = os.path.join(self.reports_dir, "html")
        
    def generate_excel_report(self, test_results, summary):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(self.excel_dir, f"Mock_Test_Execution_Report_{timestamp}.xlsx")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Execution"
        
        # Headers
        headers = ["Test Case ID", "Module", "Test Name", "Priority", "Execution Time", "Result", "Remarks", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
        # Data
        for row, result in enumerate(test_results, 2):
            ws.cell(row=row, column=1, value=result['id'])
            ws.cell(row=row, column=2, value=result['module'])
            ws.cell(row=row, column=3, value=result['name'])
            ws.cell(row=row, column=4, value=result['priority'])
            ws.cell(row=row, column=5, value=result['time'])
            ws.cell(row=row, column=6, value=result['status'])
            ws.cell(row=row, column=7, value=result['remarks'])
            ws.cell(row=row, column=8, value=result['timestamp'])
            
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Test Cases", summary['total']])
        ws_summary.append(["Executed", summary['executed']])
        ws_summary.append(["Passed", summary['passed']])
        ws_summary.append(["Failed", summary['failed']])
        ws_summary.append(["Skipped", summary['skipped']])
        ws_summary.append(["Pass Percentage", f"{summary['pass_percentage']}%"])
        
        wb.save(filename)
        return filename

    def generate_html_report(self, test_results, summary):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.join(self.html_dir, f"Mock_Test_Execution_Report_{timestamp}.html")
        
        html_content = f"""
        <html>
        <head>
            <title>Mock Test Execution Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1, h2 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; margin-bottom: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
                .passed {{ color: green; font-weight: bold; }}
                .summary-box {{ background: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <h1>Mock Test Execution Report</h1>
            <p><i>Simulated Results for Demonstration</i></p>
            
            <div class="summary-box">
                <h2>Execution Statistics</h2>
                <p>Total Test Cases: {summary['total']}</p>
                <p>Executed: {summary['executed']}</p>
                <p>Passed: {summary['passed']}</p>
                <p>Failed: {summary['failed']}</p>
                <p>Skipped: {summary['skipped']}</p>
                <p>Pass Percentage: {summary['pass_percentage']}%</p>
            </div>
            
            <h2>Test Details</h2>
            <table>
                <tr>
                    <th>ID</th><th>Module</th><th>Name</th><th>Priority</th><th>Time</th><th>Result</th><th>Remarks</th>
                </tr>
        """
        
        for r in test_results:
            status_class = "passed" if r['status'] == "Pass" else ""
            html_content += f"""
                <tr>
                    <td>{r['id']}</td>
                    <td>{r['module']}</td>
                    <td>{r['name']}</td>
                    <td>{r['priority']}</td>
                    <td>{r['time']}s</td>
                    <td class="{status_class}">{r['status']}</td>
                    <td>{r['remarks']}</td>
                </tr>
            """
            
        html_content += """
            </table>
        </body>
        </html>
        """
        
        with open(filename, "w") as f:
            f.write(html_content)
            
        return filename
